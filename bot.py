"""
Legacy Spirits - Telegram order bot (v1)

Employees send their orders to the bot in plain text, one item per line, e.g.

    2 Tito's 750
    6 Skol vodka E
    Hornitos black barrel 750 x3
    1 Jameson 1.75

The bot turns each submission into a tidy Excel order sheet and forwards it to
Mahesh, then confirms back to the employee.

v1 COLLECTS and COMPILES orders. It does NOT yet look up item codes / prices
from the ABC price list - those columns are left blank for the matching step.
(Automated price-list + DA matching is the planned v2.)

Environment variables (set these in Render - never in code or chat):
    BOT_TOKEN         - from @BotFather
    MAHESH_CHAT_ID    - your Telegram chat id (use /myid in the bot to get it)
    ALLOWED_USER_IDS  - optional, comma-separated ids allowed to order
    PORT              - set automatically by Render
"""
import os
import re
import logging
from datetime import datetime
from threading import Thread
from http.server import BaseHTTPRequestHandler, HTTPServer

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ContextTypes, filters,
)

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
log = logging.getLogger("legacy-spirits-bot")

BOT_TOKEN = os.environ.get("BOT_TOKEN")
MAHESH_CHAT_ID = os.environ.get("MAHESH_CHAT_ID")
ALLOWED_USER_IDS = {
    x.strip() for x in os.environ.get("ALLOWED_USER_IDS", "").split(",") if x.strip()
}

# ---- size handling -------------------------------------------------------
# Size codes: A=750ml, B=375ml, C=200ml, D=50ml, E=Liter, F=1.75L
SIZE_TOKENS = {
    "750": ("A", "750ml"), "750ml": ("A", "750ml"),
    "375": ("B", "375ml"), "375ml": ("B", "375ml"), "pint": ("B", "375ml"),
    "200": ("C", "200ml"), "200ml": ("C", "200ml"),
    "50": ("D", "50ml"), "50ml": ("D", "50ml"), "mini": ("D", "50ml"),
    "1l": ("E", "Liter"), "1000": ("E", "Liter"), "1000ml": ("E", "Liter"),
    "liter": ("E", "Liter"), "litre": ("E", "Liter"), "ltr": ("E", "Liter"),
    "1.75": ("F", "1.75L"), "1.75l": ("F", "1.75L"), "1750": ("F", "1.75L"),
    "1.75lt": ("F", "1.75L"), "handle": ("F", "1.75L"),
}
SINGLE_LETTER_SIZES = {
    "a": ("A", "750ml"), "b": ("B", "375ml"), "c": ("C", "200ml"),
    "d": ("D", "50ml"), "e": ("E", "Liter"), "f": ("F", "1.75L"),
}


def parse_line(line: str):
    """Parse one order line into (qty, product, size_code, size_label)."""
    raw = line.strip()
    if not raw:
        return None

    text = raw
    qty = 1

    # quantity as "x3", "3x", or a leading number
    m = re.search(r"\b[xX]\s*(\d+)\b", text)
    if m:
        qty = int(m.group(1))
        text = text[: m.start()] + " " + text[m.end():]
    elif re.search(r"\b(\d+)\s*[xX]\b", text):
        m = re.search(r"\b(\d+)\s*[xX]\b", text)
        qty = int(m.group(1))
        text = text[: m.start()] + " " + text[m.end():]
    else:
        m = re.match(r"^\s*(\d+)\s+(?=\S)", text)
        if m:
            qty = int(m.group(1))
            text = text[m.end():]

    # size: scan tokens for a known size word/number
    size_code, size_label = "", ""
    kept = []
    for tok in text.split():
        key = tok.lower().strip(".,;:()")
        if not size_code and key in SIZE_TOKENS:
            size_code, size_label = SIZE_TOKENS[key]
            continue
        kept.append(tok)

    # if still no size, allow a trailing single-letter size code (A-F)
    if not size_code and kept:
        last = kept[-1].lower().strip(".,;:()")
        if last in SINGLE_LETTER_SIZES:
            size_code, size_label = SINGLE_LETTER_SIZES[last]
            kept = kept[:-1]

    product = " ".join(kept).strip(" -,")
    if not product:
        return None
    return qty, product, size_code, size_label


# ---- excel ---------------------------------------------------------------
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)

HEADERS = ["#", "Code", "Product (as ordered)", "Size", "Qty",
           "Regular Price", "DA Sale Price", "Savings"]


def build_excel(employee: str, items: list) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    ws["A1"] = "Legacy Spirits - Employee Order"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"From: {employee}    |    {datetime.now():%Y-%m-%d %H:%M}"
    ws["A2"].font = Font(italic=True, color="555555")
    # row 3 left blank as a spacer

    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for i, (qty, product, size_code, size_label) in enumerate(items, start=1):
        r = 4 + i
        size_disp = f"{size_code} ({size_label})" if size_code else (size_label or "")
        values = [i, "", product, size_disp, qty, "", "", ""]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            if c in (1, 4, 5):
                cell.alignment = Alignment(horizontal="center")

    widths = [5, 12, 38, 14, 6, 14, 14, 12]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A5"

    safe = re.sub(r"[^A-Za-z0-9]+", "_", employee).strip("_") or "order"
    path = os.path.join("/tmp", f"order_{safe}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    wb.save(path)
    return path


# ---- handlers ------------------------------------------------------------
def who(update: Update) -> str:
    u = update.effective_user
    name = u.full_name if u else "Unknown"
    uname = f"@{u.username}" if u and u.username else ""
    return (name + (f" ({uname})" if uname else "")).strip()


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Legacy Spirits order bot\n\n"
        "Send your order, one item per line, like:\n"
        "  2 Tito's 750\n"
        "  6 Skol vodka E\n"
        "  Hornitos black barrel 750 x3\n\n"
        "I'll send the order sheet straight to Mahesh.\n"
        "Size can be 750 / 375 / 200 / 50, liter, 1.75 - or codes A-F.\n\n"
        "Type /myid to see your Telegram id."
    )


async def myid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Your chat id: {update.effective_chat.id}")


async def handle_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    if ALLOWED_USER_IDS and user_id not in ALLOWED_USER_IDS:
        await update.message.reply_text(
            "Sorry, you're not on the orders list yet. Ask Mahesh to add you."
        )
        return

    text = update.message.text or ""
    items, skipped = [], []
    for line in text.splitlines():
        if not line.strip():
            continue
        parsed = parse_line(line)
        (items if parsed else skipped).append(parsed if parsed else line.strip())

    if not items:
        await update.message.reply_text(
            "I couldn't read any items. Send one item per line, e.g. `2 Tito's 750`."
        )
        return

    employee = who(update)
    path = build_excel(employee, items)

    if not MAHESH_CHAT_ID:
        await update.message.reply_text(
            "Order received, but delivery isn't set up yet (MAHESH_CHAT_ID missing).\n"
            "Mahesh: message me, run /myid, then add that id in Render."
        )
        return

    caption = f"New order from {employee} - {len(items)} item(s)"
    with open(path, "rb") as f:
        await context.bot.send_document(
            chat_id=MAHESH_CHAT_ID, document=f,
            filename=os.path.basename(path), caption=caption,
        )

    msg = f"Sent your order to Mahesh - {len(items)} item(s)."
    if skipped:
        msg += "\nCouldn't read: " + "; ".join(skipped[:5])
    await update.message.reply_text(msg)


# ---- keep-alive web server (so Render's free web service has a port) ------
class _Health(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"ok")

    def log_message(self, *args):
        pass


def start_health_server():
    port = int(os.environ.get("PORT", "10000"))
    HTTPServer(("0.0.0.0", port), _Health).serve_forever()


def main():
    if not BOT_TOKEN:
        raise SystemExit("BOT_TOKEN env var is not set.")
    Thread(target=start_health_server, daemon=True).start()
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("myid", myid))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_order))
    log.info("Bot starting (polling)...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
